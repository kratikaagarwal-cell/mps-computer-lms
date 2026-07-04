from flask import Flask, render_template, render_template_string, request, redirect, session, send_from_directory, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_caching import Cache
from datetime import datetime
import openpyxl
import os
from supabase import create_client

app = Flask(__name__)

# ── Secrets: always loaded from environment variables ──────────────────────────
# Never put real values here. Set them in your hosting platform's env settings.
app.secret_key = os.environ.get("SECRET_KEY", "change-this-in-production")

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise RuntimeError("SUPABASE_URL and SUPABASE_KEY environment variables must be set.")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL environment variable must be set.")

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL

# ── Database connection pool tuned for Render + Supabase FREE TIER ─────────────
# IMPORTANT: each gunicorn worker process gets its OWN pool. Total possible
# connections = workers * (pool_size + max_overflow). With 2 workers this
# gives a max of 2 * (5+5) = 20 connections - safely under Supabase's free
# tier connection cap. Do NOT raise this without also raising your Supabase
# plan / switching to the pooler connection string (see note below).
#
# Also: make sure DATABASE_URL uses Supabase's "Transaction" pooler
# (port 6543), not the direct connection (port 5432) - Project Settings ->
# Database -> Connection string -> Transaction pooler. The pooler is built
# to handle many short-lived connections from apps like this one; the
# direct connection has a much lower hard limit.
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,       # drop stale connections automatically
    "pool_size": 5,              # keep 5 connections open per worker
    "max_overflow": 5,           # allow up to 5 extra per worker at peak
    "pool_timeout": 30,          # wait max 30s for a free connection
    "pool_recycle": 1800,        # recycle connections every 30 min
}
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

# ── Simple in-process cache (upgrade to Redis for multi-server deployments) ────
app.config["CACHE_TYPE"] = "SimpleCache"
app.config["CACHE_DEFAULT_TIMEOUT"] = 60   # 60-second TTL for dashboard queries
cache = Cache(app)

db = SQLAlchemy(app)

def _student_cache_key(student_id):
    """Cache key scoped per student so invalidation is surgical."""
    return f"student_dashboard_{student_id}"

def invalidate_student_cache(student_id):
    """Call this whenever a student's progress/data changes."""
    cache.delete(_student_cache_key(student_id))

# ─────────────────────────────────────────────
# MODELS
# ─────────────────────────────────────────────

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True)
    password = db.Column(db.String(50))

class Teacher(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True)
    password = db.Column(db.String(50))
    name = db.Column(db.String(100))
    subject = db.Column(db.String(100), default="Computer Science")

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    admission_no = db.Column(db.String(50), unique=True)
    student_name = db.Column(db.String(100))
    dob = db.Column(db.String(20))
    student_class = db.Column(db.String(10))
    section = db.Column(db.String(10))
    roll_no = db.Column(db.String(10))
    parent_mobile = db.Column(db.String(15))
    status = db.Column(db.String(20), default="Active")

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200))
    original_name = db.Column(db.String(200))
    chapter = db.Column(db.String(100))
    chapter_no = db.Column(db.Integer, default=0)
    student_class = db.Column(db.String(10), default="6")
    section = db.Column(db.String(20))
    description = db.Column(db.String(300))
    file_type = db.Column(db.String(20), default="notes")
    # FIX #3: support external URL for books
    external_url = db.Column(db.String(500), default="")
    uploaded_at = db.Column(db.String(30), default="")

class Assignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200))
    chapter = db.Column(db.String(100))
    chapter_no = db.Column(db.Integer, default=0)
    student_class = db.Column(db.String(10), default="6")
    section = db.Column(db.String(20))
    description = db.Column(db.Text)
    due_date = db.Column(db.String(30))
    created_at = db.Column(db.String(30), default="")
    filename = db.Column(db.String(200), default="")

class ChapterProgress(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"))
    chapter_no = db.Column(db.Integer)
    student_class = db.Column(db.String(10), default="6")
    video_done = db.Column(db.Boolean, default=False)
    notes_done = db.Column(db.Boolean, default=False)
    book_done = db.Column(db.Boolean, default=False)
    quiz_done = db.Column(db.Boolean, default=False)
    game_done = db.Column(db.Boolean, default=False)
    # FIX #9: renamed feedback_done -> query_done (kept column name for DB compat)
    feedback_done = db.Column(db.Boolean, default=False)
    updated_at = db.Column(db.String(30), default="")

class AssignmentSubmission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"))
    assignment_id = db.Column(db.Integer, db.ForeignKey("assignment.id"))
    status = db.Column(db.String(20), default="pending")
    submitted_at = db.Column(db.String(30), default="")
    remarks = db.Column(db.String(300), default="")

class Quiz(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200))
    chapter = db.Column(db.String(100))
    chapter_no = db.Column(db.Integer, default=0)
    student_class = db.Column(db.String(10), default="6")
    section = db.Column(db.String(20))
    description = db.Column(db.String(300), default="")
    quiz_type = db.Column(db.String(20), default="mcq")
    external_link = db.Column(db.String(500), default="")
    created_at = db.Column(db.String(30), default="")
    questions = db.relationship("QuizQuestion", backref="quiz", cascade="all, delete-orphan")
    attempts = db.relationship("QuizAttempt", backref="quiz", cascade="all, delete-orphan")

class QuizQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey("quiz.id"))
    question_text = db.Column(db.Text)
    option_a = db.Column(db.String(300))
    option_b = db.Column(db.String(300))
    option_c = db.Column(db.String(300))
    option_d = db.Column(db.String(300))
    correct_option = db.Column(db.String(1))

class QuizAttempt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"))
    quiz_id = db.Column(db.Integer, db.ForeignKey("quiz.id"))
    score = db.Column(db.Integer, default=0)
    total = db.Column(db.Integer, default=0)
    attempted_at = db.Column(db.String(30), default="")
    student = db.relationship("Student", backref="quiz_attempts")

class Game(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200))
    chapter = db.Column(db.String(100))
    chapter_no = db.Column(db.Integer, default=0)
    student_class = db.Column(db.String(10), default="6")
    section = db.Column(db.String(20))
    created_at = db.Column(db.String(30), default="")
    game_type = db.Column(db.String(30), default="matching")
    pairs = db.relationship("GamePair", backref="game", cascade="all, delete-orphan")
    scores = db.relationship("GameScore", backref="game", cascade="all, delete-orphan")

class GamePair(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    game_id = db.Column(db.Integer, db.ForeignKey("game.id"))
    term = db.Column(db.String(200))
    definition = db.Column(db.String(200))

    def to_dict(self):
        return {"id": self.id, "term": self.term, "definition": self.definition}

class GameScore(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"))
    game_id = db.Column(db.Integer, db.ForeignKey("game.id"))
    score = db.Column(db.Integer, default=0)
    time_seconds = db.Column(db.Integer, default=999)
    played_at = db.Column(db.String(30), default="")
    student = db.relationship("Student", backref="game_scores")

class ChapterVideo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    chapter_no = db.Column(db.Integer)
    chapter = db.Column(db.String(100))
    student_class = db.Column(db.String(10), default="6")
    section = db.Column(db.String(20))
    youtube_url = db.Column(db.String(500))
    title = db.Column(db.String(200))
    added_at = db.Column(db.String(30), default="")

class ChapterConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    chapter_no = db.Column(db.Integer)
    student_class = db.Column(db.String(10), default="6")
    is_active = db.Column(db.Boolean, default=False)
    active_month = db.Column(db.String(20), default="")
    has_video = db.Column(db.Boolean, default=False)
    has_notes = db.Column(db.Boolean, default=False)
    has_book = db.Column(db.Boolean, default=False)
    has_quiz = db.Column(db.Boolean, default=False)
    has_game = db.Column(db.Boolean, default=False)
    # FIX #9: has_feedback renamed semantically to has_query but kept DB column
    has_feedback = db.Column(db.Boolean, default=True)

# FIX #9: New Query model (replaces Feedback as progress item)
class StudentQuery(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"))
    chapter_no = db.Column(db.Integer)
    student_class = db.Column(db.String(10), default="6")
    comment = db.Column(db.Text, default="")
    submitted_at = db.Column(db.String(30), default="")
    # FIX #9: teacher reply field
    teacher_reply = db.Column(db.Text, default="")
    replied_at = db.Column(db.String(30), default="")
    student = db.relationship("Student", backref="queries")

class Feedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"))
    chapter_no = db.Column(db.Integer)
    student_class = db.Column(db.String(10), default="6")
    rating = db.Column(db.Integer, default=5)
    comment = db.Column(db.Text, default="")
    submitted_at = db.Column(db.String(30), default="")
    student = db.relationship("Student", backref="feedbacks")

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

ALL_CHAPTERS = {
    "6": [
        {"no": 1, "name": "Internet Services"},
        {"no": 2, "name": "More on Excel"},
        {"no": 3, "name": "Formulas & Functions"},
        {"no": 4, "name": "Canva"},
        {"no": 5, "name": "Introduction to GIMP"},
        {"no": 6, "name": "Present Yourself Online"},
        {"no": 7, "name": "Introduction to Python"},
        {"no": 8, "name": "Introduction to Data Science"},
        {"no": 9, "name": "Robots & Sensors"},
    ],
    "7": [
        {"no": 1,  "name": "Advanced Features of Excel"},
        {"no": 2,  "name": "More on GIMP"},
        {"no": 3,  "name": "Digital Footprints"},
        {"no": 4,  "name": "Introduction to App Development"},
        {"no": 5,  "name": "Introduction to HTML and CSS"},
        {"no": 6,  "name": "More on CSS"},
        {"no": 7,  "name": "More on Python"},
        {"no": 8,  "name": "Computer Vision"},
        {"no": 9,  "name": "AI in Robotics"},
    ],
    "8": [
        {"no": 1,  "name": "OpenShot Video Editor"},
        {"no": 2,  "name": "Database Management System"},
        {"no": 3,  "name": "Structured Query Language"},
        {"no": 4,  "name": "Fear of Missing Out"},
        {"no": 5,  "name": "Lists and Tables in HTML"},
        {"no": 6,  "name": "Links & Frames in HTML"},
        {"no": 7,  "name": "Advanced Python"},
        {"no": 8,  "name": "Natural Language Processing"},
        {"no": 9,  "name": "Next Generation of Robots"},
    ],
    "9": [
        {"no": 1,  "name": "PA U1 - Communication Skills-1"},
        {"no": 2,  "name": "PA U2 - Self-Management Skills-1"},
        {"no": 3,  "name": "PA U3 - ICT Skills-1"},
        {"no": 4,  "name": "PA U4 - Entrepreneurial Skills-1"},
        {"no": 5,  "name": "PA U5 - Green Skills-1"},
        {"no": 6,  "name": "PB U1 - Introduction to IT-ITeS Industry"},
        {"no": 7,  "name": "PB U2 - Data Entry & Keyboarding Skills"},
        {"no": 8,  "name": "PB U3 - Digital Documentation"},
        {"no": 9,  "name": "PB U4 - Electronic Spreadsheet"},
        {"no": 10, "name": "PB U5 - Digital Presentation"},
    ],
}

ALL_SECTIONS = {
    "6": ["6A","6B","6C","6D","6E","6F","6G"],
    "7": ["7A","7B","7C","7D","7E","7F","7G"],
    "8": ["8A","8B","8C","8D","8E","8F","8G"],
    "9": ["9A","9B","9C","9D","9E","9F","9G","9H"],
}

MONTHS = ["January","February","March","April","May","June","July","August","September","October","November","December"]

# Legacy alias for old code paths
CHAPTERS = ALL_CHAPTERS["6"]
SECTIONS  = ALL_SECTIONS["6"]

def normalize_section(student_class, section_value):
    """The rest of the app expects Student.section in the combined
    'class+section' form used by ALL_SECTIONS, e.g. '6A' — not just 'A'.
    If someone enters (or imports) a bare section letter/number, prefix it
    with the student's class so dropdowns, bulk-delete-by-section, and
    notes/books visibility all match correctly."""
    cls = (student_class or "").strip()
    sec = (section_value or "").strip().upper()
    if not sec:
        return sec
    if sec.startswith(cls):
        return sec
    return f"{cls}{sec}"

def now_str():
    return datetime.now().strftime("%d %b %Y, %I:%M %p")

def safe_filename(filename):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_")
    return ts + filename.replace(" ", "_")

def get_youtube_embed(url):
    if not url: return ""
    if "embed" in url: return url
    if "youtu.be/" in url:
        vid = url.split("youtu.be/")[-1].split("?")[0]
    elif "v=" in url:
        vid = url.split("v=")[-1].split("&")[0]
    else:
        return url
    return f"https://www.youtube.com/embed/{vid}"

def to_embeddable_url(url):
    """Google Drive 'view' share links (.../file/d/<id>/view) block being
    embedded in an iframe. Convert them to the /preview form, which Google
    allows to be embedded, so the read-only viewer can actually show them."""
    if "drive.google.com" in url and "/file/d/" in url:
        try:
            file_id = url.split("/file/d/")[1].split("/")[0]
            return f"https://drive.google.com/file/d/{file_id}/preview"
        except IndexError:
            return url
    return url

def split_sections(section_value):
    """A Note.section can hold one section, a comma-joined list ('6A,6C,6F'),
    or 'All Sections'. Always return a clean list of individual sections."""
    if not section_value:
        return []
    return [s.strip() for s in section_value.split(",") if s.strip()]

def note_visible_to_section(note_section, student_sec):
    parts = split_sections(note_section)
    return student_sec in parts or "All Sections" in parts

class NoteGroup:
    """Display-only wrapper: one row per unique (file, chapter) combo, with
    every section merged into one badge list. Handles both a single upload
    row whose section field already holds several comma-joined sections,
    and (for any older data) multiple rows that still need merging."""
    def __init__(self, note):
        self.ids = [note.id]
        self.filename = note.filename
        self.original_name = note.original_name
        self.chapter = note.chapter
        self.chapter_no = note.chapter_no
        self.sections = split_sections(note.section)

    def add(self, note):
        self.ids.append(note.id)
        for s in split_sections(note.section):
            if s not in self.sections:
                self.sections.append(s)

    @property
    def id_csv(self):
        return ",".join(str(i) for i in self.ids)

    @property
    def section_display(self):
        return ", ".join(self.sections) if self.sections else "All Sections"

def group_notes_by_file(note_list):
    """Collapse Note rows that share the same file/URL + chapter (uploaded
    to multiple sections) into a single row for cleaner teacher-side display."""
    groups = {}
    order = []
    for n in note_list:
        key = (n.filename, n.chapter_no)
        if key in groups:
            groups[key].add(n)
        else:
            groups[key] = NoteGroup(n)
            order.append(key)
    return [groups[k] for k in order]

def get_chapter_progress_pct(prog, config):
    if not prog or not config: return 0
    total = done = 0
    if config.has_video:    total += 1; done += int(prog.video_done)
    if config.has_notes:    total += 1; done += int(prog.notes_done)
    if config.has_book:     total += 1; done += int(prog.book_done)
    if config.has_quiz:     total += 1; done += int(prog.quiz_done)
    if config.has_game:     total += 1; done += int(prog.game_done)
    # FIX #9: feedback_done no longer counts toward progress
    return int((done / total) * 100) if total else 0

# ─────────────────────────────────────────────
# HOME
# ─────────────────────────────────────────────

@app.route("/")
def home():
    return render_template("home.html")

# ─────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        admin = Admin.query.filter_by(username=request.form["username"], password=request.form["password"]).first()
        if admin:
            session["admin"] = admin.username
            return redirect("/admin")
        return render_template("login.html", error="Invalid credentials!")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

@app.route("/teacher_login", methods=["GET","POST"])
def teacher_login():
    if request.method == "POST":
        teacher = Teacher.query.filter_by(username=request.form["username"], password=request.form["password"]).first()
        if teacher:
            session["teacher"] = teacher.username
            session["teacher_name"] = teacher.name
            return redirect("/teacher")
        return render_template("teacher_login.html", error="Invalid credentials!")
    return render_template("teacher_login.html")

@app.route("/teacher_logout")
def teacher_logout():
    session.pop("teacher", None)
    session.pop("teacher_name", None)
    return redirect("/")

# ─────────────────────────────────────────────
# ADMIN
# ─────────────────────────────────────────────

@app.route("/admin")
def admin():
    if "admin" not in session: return redirect("/login")
    total_students = Student.query.count()
    total_notes    = Note.query.filter_by(file_type="notes").count()
    total_books    = Note.query.filter_by(file_type="book").count()
    total_quizzes  = Quiz.query.count()
    total_assignments = Assignment.query.count()
    recent_notes   = Note.query.order_by(Note.id.desc()).limit(5).all()
    recent_assignments = Assignment.query.order_by(Assignment.id.desc()).limit(5).all()
    all_sections = ["6A","6B","6C","6D","6E","6F","6G",
                    "7A","7B","7C","7D","7E","7F","7G",
                    "8A","8B","8C","8D","8E","8F","8G",
                    "9A","9B","9C","9D","9E","9F","9G","9H"]
    section_counts = {sec: Student.query.filter_by(section=sec).count() for sec in all_sections}
    return render_template("admin.html",
        total_students=total_students, total_notes=total_notes,
        total_books=total_books, total_quizzes=total_quizzes,
        total_assignments=total_assignments, recent_notes=recent_notes,
        recent_assignments=recent_assignments, section_counts=section_counts,
        sections=all_sections)

@app.route("/students")
def students():
    if "admin" not in session: return redirect("/login")
    search  = request.args.get("search","")
    section = request.args.get("section","")
    cls     = request.args.get("class","")
    data = Student.query
    if search:  data = data.filter((Student.student_name.contains(search))|(Student.admission_no.contains(search)))
    if section: data = data.filter_by(section=section)
    if cls:     data = data.filter_by(student_class=cls)
    all_sections = ["6A","6B","6C","6D","6E","6F","6G",
                    "7A","7B","7C","7D","7E","7F","7G",
                    "8A","8B","8C","8D","8E","8F","8G",
                    "9A","9B","9C","9D","9E","9F","9G","9H"]
    return render_template("students.html", students=data.all(), total=Student.query.count(),
        sections=all_sections, search=search, section=section)

FIX_SECTIONS_RESULT_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Sections Fixed — MPS LMS</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Segoe UI', Arial, sans-serif; background: linear-gradient(135deg, #0f4c81, #1abc9c); min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
.box { background: white; border-radius: 24px; padding: 48px 36px; text-align: center; box-shadow: 0 20px 60px rgba(0,0,0,0.2); max-width: 440px; width: 100%; }
.icon { font-size: 70px; display: block; margin-bottom: 16px; }
h2 { font-size: 26px; font-weight: 800; color: #0f4c81; margin-bottom: 8px; }
p { font-size: 14px; color: #888; margin-bottom: 6px; }
.num { font-size: 42px; font-weight: 800; color: #1abc9c; margin: 16px 0 6px; }
.note { background: #f0f4f8; border-radius: 10px; padding: 10px 14px; font-size: 13px; color: #555; margin: 18px 0 24px; line-height: 1.5; }
.btn { display: inline-block; padding: 12px 28px; border-radius: 25px; font-size: 14px; font-weight: 700; color: white; text-decoration: none; background: linear-gradient(135deg, #0f4c81, #1abc9c); margin: 0 6px; }
</style>
</head>
<body>
<div class="box">
  <span class="icon">🔧</span>
  <h2>Section Values Fixed!</h2>
  <p>Student sections have been normalized.</p>
  <div class="num">{{ fixed }}</div>
  <p>record{{ 's' if fixed != 1 else '' }} updated</p>
  {% if fixed > 0 %}
  <div class="note">Bare section letters (e.g. "C") were rewritten to the full class+section code (e.g. "6C") so bulk delete-by-section now matches correctly.</div>
  {% else %}
  <div class="note">Everything was already in the correct format — no changes were needed.</div>
  {% endif %}
  <a href="/students" class="btn">👨‍🎓 View Students</a>
  <a href="/admin" class="btn">📊 Dashboard</a>
</div>
</body>
</html>
"""

@app.route("/fix_student_sections")
def fix_student_sections():
    # One-click cleanup for students whose Section was entered as just
    # 'A' etc. instead of the combined 'class+section' form ('6A') the
    # rest of the app expects — which is why bulk delete-by-section
    # (and other section matching) was silently finding 0 students.
    if "admin" not in session:
        return redirect("/login")
    fixed = 0
    for s in Student.query.all():
        correct = normalize_section(s.student_class, s.section)
        if correct != s.section:
            s.section = correct
            fixed += 1
    db.session.commit()
    return render_template_string(FIX_SECTIONS_RESULT_HTML, fixed=fixed)

@app.route("/delete/<int:id>")
def delete_student(id):
    if "admin" not in session: 
        return redirect("/login")
    
    # Step 1: Verify student exists
    student = db.session.get(Student, id)
    if not student:
        return redirect("/students")
    
    # Step 2: Delete all related records BEFORE deleting the student
    # This prevents Foreign Key constraint violations
    ChapterProgress.query.filter_by(student_id=id).delete(synchronize_session=False)
    AssignmentSubmission.query.filter_by(student_id=id).delete(synchronize_session=False)
    QuizAttempt.query.filter_by(student_id=id).delete(synchronize_session=False)
    GameScore.query.filter_by(student_id=id).delete(synchronize_session=False)
    StudentQuery.query.filter_by(student_id=id).delete(synchronize_session=False)
    Feedback.query.filter_by(student_id=id).delete(synchronize_session=False)
    
    # Step 3: Now delete the student record
    db.session.delete(student)
    db.session.commit()
    
    return redirect("/students")

# ── Chapter-wise progress report (teacher + admin) ────────────────────────────
@app.route("/progress_report")
def progress_report():
    if "teacher" not in session and "admin" not in session:
        return redirect("/teacher_login")
    sel_class   = request.args.get("class", "6")
    sel_section = request.args.get("section", "")
    CHAPTERS    = ALL_CHAPTERS.get(sel_class, ALL_CHAPTERS["6"])
    sections    = ALL_SECTIONS.get(sel_class, [])
    configs     = {c.chapter_no: c for c in ChapterConfig.query.filter_by(student_class=sel_class).all()}

    q = Student.query.filter_by(student_class=sel_class)
    if sel_section:
        q = q.filter_by(section=sel_section)
    students = q.order_by(Student.section, Student.roll_no).all()

    # Build progress map: {student_id: {chapter_no: pct}}
    all_progress = ChapterProgress.query.filter_by(student_class=sel_class).all()
    prog_map = {}
    for p in all_progress:
        prog_map.setdefault(p.student_id, {})[p.chapter_no] = p

    # Build chapter summary: {chapter_no: {completed, in_progress, not_started, total}}
    chapter_summary = {}
    for ch in CHAPTERS:
        cno = ch["no"]
        cfg = configs.get(cno)
        completed = in_progress = 0
        for s in students:
            pct = get_chapter_progress_pct(prog_map.get(s.id, {}).get(cno), cfg)
            if pct == 100: completed += 1
            elif pct > 0:  in_progress += 1
        chapter_summary[cno] = {
            "completed":   completed,
            "in_progress": in_progress,
            "not_started": len(students) - completed - in_progress,
            "total":       len(students)
        }

    # Build student rows: [{student, chapter_pcts: {cno: pct}, overall_pct}]
    student_rows = []
    for s in students:
        pcts = {}
        for ch in CHAPTERS:
            cno = ch["no"]
            pcts[cno] = get_chapter_progress_pct(prog_map.get(s.id, {}).get(cno), configs.get(cno))
        overall = int(sum(pcts.values()) / len(CHAPTERS)) if CHAPTERS else 0
        student_rows.append({"student": s, "pcts": pcts, "overall": overall})

    return render_template("progress_report.html",
        chapters=CHAPTERS, chapter_summary=chapter_summary,
        student_rows=student_rows, sel_class=sel_class,
        sel_section=sel_section, sections=sections,
        total_students=len(students))

# ── Printable single-student report ───────────────────────────────────────────
@app.route("/student_report/<int:student_id>")
def student_report(student_id):
    if "teacher" not in session and "admin" not in session:
        return redirect("/teacher_login")
    student  = db.session.get(Student, student_id)
    if not student: return "Student not found", 404
    cls      = student.student_class
    CHAPTERS = ALL_CHAPTERS.get(cls, ALL_CHAPTERS["6"])
    configs  = {c.chapter_no: c for c in ChapterConfig.query.filter_by(student_class=cls).all()}
    progs    = {p.chapter_no: p for p in ChapterProgress.query.filter_by(student_id=student_id, student_class=cls).all()}

    chapter_data = []
    for ch in CHAPTERS:
        cno  = ch["no"]
        cfg  = configs.get(cno)
        prog = progs.get(cno)
        pct  = get_chapter_progress_pct(prog, cfg)
        chapter_data.append({
            "no":     cno,
            "name":   ch["name"],
            "pct":    pct,
            "video":  prog.video_done  if prog else False,
            "notes":  prog.notes_done  if prog else False,
            "book":   prog.book_done   if prog else False,
            "quiz":   prog.quiz_done   if prog else False,
            "game":   prog.game_done   if prog else False,
            "has_video": cfg.has_video if cfg else False,
            "has_notes": cfg.has_notes if cfg else False,
            "has_book":  cfg.has_book  if cfg else False,
            "has_quiz":  cfg.has_quiz  if cfg else False,
            "has_game":  cfg.has_game  if cfg else False,
        })
    overall = int(sum(c["pct"] for c in chapter_data) / len(CHAPTERS)) if CHAPTERS else 0
    return render_template("student_report.html",
        student=student, chapter_data=chapter_data, overall=overall,
        generated_on=now_str())

@app.route("/delete_section", methods=["POST"])
def delete_section():
    if "admin" not in session: return redirect("/login")
    sec = request.form.get("del_section", "").strip()
    if not sec:
        return redirect("/students")
    student_ids = [s.id for s in Student.query.filter_by(section=sec).with_entities(Student.id).all()]
    if student_ids:
        ChapterProgress.query.filter(ChapterProgress.student_id.in_(student_ids)).delete(synchronize_session=False)
        AssignmentSubmission.query.filter(AssignmentSubmission.student_id.in_(student_ids)).delete(synchronize_session=False)
        QuizAttempt.query.filter(QuizAttempt.student_id.in_(student_ids)).delete(synchronize_session=False)
        GameScore.query.filter(GameScore.student_id.in_(student_ids)).delete(synchronize_session=False)
        StudentQuery.query.filter(StudentQuery.student_id.in_(student_ids)).delete(synchronize_session=False)
        Feedback.query.filter(Feedback.student_id.in_(student_ids)).delete(synchronize_session=False)
        Student.query.filter(Student.id.in_(student_ids)).delete(synchronize_session=False)
        db.session.commit()
    return redirect("/students")

@app.route("/download_student_template")
def download_student_template():
    # Blank Excel with the exact required headers (+ one greyed-out example
    # row) so teachers can't typo a column name when importing students.
    if "admin" not in session:
        return redirect("/login")

    from io import BytesIO
    from flask import send_file
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["Admission_No", "Student_Name", "DOB", "Class", "Section", "Roll_No", "Parent_Mobile"]
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # One example row, styled grey/italic so it's obviously a sample to replace
    example = ["100234", "Aarav Sharma", "12-05-2014", "6", "6A", "12", "9876543210"]
    sample_font = Font(italic=True, color="999999", name="Arial", size=10)
    for col, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = sample_font

    widths = [16, 24, 14, 10, 12, 10, 16]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="student_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/upload_students", methods=["GET","POST"])
def upload_students():

    if "admin" not in session:
        return redirect("/login")

    if request.method == "POST":

        file = request.files.get("file")

        if not file:
            return "No file selected"

        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

        path = os.path.join(
            app.config["UPLOAD_FOLDER"],
            file.filename
        )

        file.save(path)

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        headers = [
            str(cell.value).strip()
            for cell in ws[1]
        ]

        existing_admissions = {
            s.admission_no 
            for s in Student.query.with_entities(Student.admission_no).all()
        }

        students_to_add = []

        for row in ws.iter_rows(min_row=2, values_only=True):

            data = dict(zip(headers,row))

            admission = str(
                data.get("Admission_No","")
            ).strip()


            if admission and admission not in existing_admissions:

                cls = str(data.get("Class",""))
                students_to_add.append(
                    Student(
                        admission_no=admission,
                        student_name=str(data.get("Student_Name","")),
                        dob=str(data.get("DOB","")),
                        student_class=cls,
                        section=normalize_section(cls, data.get("Section","")),
                        roll_no=str(data.get("Roll_No","")),
                        parent_mobile=str(data.get("Parent_Mobile",""))
                    )
                )


        db.session.bulk_save_objects(students_to_add)
        db.session.commit()


        return render_template(
            "success.html",
            total=len(students_to_add),
            filename=file.filename
        )


    return render_template("upload.html")

# ─────────────────────────────────────────────
# TEACHER
# ─────────────────────────────────────────────

@app.route("/teacher")
def teacher():
    if "teacher" not in session and "admin" not in session:
        return redirect("/teacher_login")
    sel_class = request.args.get("class","6")
    chapters  = ALL_CHAPTERS.get(sel_class, ALL_CHAPTERS["6"])
    sections  = ALL_SECTIONS.get(sel_class, ALL_SECTIONS["6"])
    notes      = Note.query.filter_by(file_type="notes", student_class=sel_class).order_by(Note.id.desc()).all()
    books      = Note.query.filter_by(file_type="book",  student_class=sel_class).order_by(Note.id.desc()).all()
    notes_grouped = group_notes_by_file(notes)
    books_grouped = group_notes_by_file(books)
    assignments= Assignment.query.filter_by(student_class=sel_class).order_by(Assignment.id.desc()).all()
    quizzes    = Quiz.query.filter_by(student_class=sel_class).order_by(Quiz.id.desc()).all()
    games      = Game.query.filter_by(student_class=sel_class).order_by(Game.id.desc()).all()
    videos     = ChapterVideo.query.filter_by(student_class=sel_class).order_by(ChapterVideo.id.desc()).all()
    configs    = {c.chapter_no: c for c in ChapterConfig.query.filter_by(student_class=sel_class).all()}
    # FIX #5/#9: load queries (not feedbacks) for teacher view
    queries    = StudentQuery.query.filter_by(student_class=sel_class).order_by(StudentQuery.id.desc()).limit(20).all()
    return render_template("teacher.html",
        notes=notes, books=books, notes_grouped=notes_grouped, books_grouped=books_grouped,
        assignments=assignments,
        quizzes=quizzes, games=games, videos=videos,
        configs=configs, chapters=chapters, sections=sections,
        months=MONTHS, teacher_name=session.get("teacher_name","Teacher"),
        total_students=Student.query.count(), queries=queries,
        sel_class=sel_class, all_classes=["6","7","8","9"])

@app.route("/upload_notes", methods=["POST"])
def upload_notes():
    if "teacher" not in session and "admin" not in session: 
        return redirect("/teacher_login")
    
    file = request.files.get("file")
    chapter = request.form["chapter"]
    chapter_no = int(request.form.get("chapter_no", 0))
    sel_class = request.form.get("student_class","6")
    
    sections_selected = request.form.getlist("section")
    if not sections_selected:
        sections_selected = [request.form.get("section", "All Sections")]
    
    description = request.form.get("description","")
    file_type = request.form.get("file_type","notes")
    external_url = request.form.get("external_url", "").strip()

    # FIX: Initialize variables
    public_url = ""
    original_filename = ""
    
    # Handle file upload
    if file and file.filename:
        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
        safe_name = safe_filename(file.filename)
        file_bytes = file.read()
        original_filename = file.filename

        import mimetypes
        mime_type, _ = mimetypes.guess_type(file.filename)
        mime_type = mime_type or "application/octet-stream"

        try:
            # Upload to Supabase with correct content-type so browser can render PDF
            supabase.storage.from_("lms-files").upload(
                safe_name,
                file_bytes,
                {"content-type": mime_type}
            )
            public_url = safe_name
            print(f"✅ File uploaded to Supabase: {safe_name} ({mime_type})")
            
        except Exception as e:
            print(f"❌ Supabase upload error: {e}")
            return redirect(f"/teacher?class={sel_class}&error=upload_failed")
    
    # Handle external URL (for books/resources)
    if external_url and not public_url:
        public_url = external_url

    # Validate we have a URL
    if not public_url:
        print(f"❌ No URL generated")
        return redirect(f"/teacher?class={sel_class}&error=no_file_or_url")

    # Create ONE note record covering every selected section — sections are
    # stored comma-joined (e.g. "6A,6C,6F") so a single upload no longer
    # multiplies into duplicate database rows.
    section_value = ",".join(sections_selected) if sections_selected else "All Sections"
    try:
        note = Note(
            filename=public_url,  # ✅ Full URL stored here
            original_name=original_filename,
            chapter=chapter,
            chapter_no=chapter_no,
            student_class=sel_class,
            section=section_value,
            description=description,
            file_type=file_type,
            external_url=external_url,
            uploaded_at=now_str()
        )
        db.session.add(note)
        print(f"✅ Note created for sections: {section_value}")
    except Exception as e:
        print(f"❌ Database error: {e}")
        db.session.rollback()
        return redirect(f"/teacher?class={sel_class}&error=db_failed")

    db.session.commit()

    # Update chapter config
    cfg = ChapterConfig.query.filter_by(chapter_no=chapter_no, student_class=sel_class).first()
    if cfg:
        if file_type == "notes": cfg.has_notes = True
        if file_type == "book": cfg.has_book = True
        db.session.commit()
    
    print(f"✅ Upload complete! URL: {public_url}")
    return redirect(f"/teacher?class={sel_class}")

@app.route("/delete_note/<int:id>")
def delete_note(id):
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    note = db.session.get(Note, id)
    sel_class = note.student_class if note else "6"
    if note:
        if note.filename:
            fpath = os.path.join(app.config["UPLOAD_FOLDER"], note.filename)
            if os.path.exists(fpath): os.remove(fpath)
        db.session.delete(note); db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/delete_note_group/<ids>")
def delete_note_group(ids):
    # ids is a comma-separated list of Note.id — deletes every section-row
    # belonging to one merged (file + chapter) entry shown on the teacher page.
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    sel_class = "6"
    for raw_id in ids.split(","):
        raw_id = raw_id.strip()
        if not raw_id.isdigit(): continue
        note = db.session.get(Note, int(raw_id))
        if note:
            sel_class = note.student_class
            if note.filename:
                fpath = os.path.join(app.config["UPLOAD_FOLDER"], note.filename)
                if os.path.exists(fpath): os.remove(fpath)
            db.session.delete(note)
    db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/add_video", methods=["POST"])
def add_video():
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    chapter_no = int(request.form["chapter_no"])
    chapter    = request.form["chapter"]
    sel_class  = request.form.get("student_class","6")
    section    = request.form["section"]
    youtube_url= request.form["youtube_url"]
    title      = request.form.get("title","Chapter Video")
    existing   = ChapterVideo.query.filter_by(chapter_no=chapter_no, student_class=sel_class, section=section).first()
    if existing:
        existing.youtube_url = youtube_url; existing.title = title; existing.added_at = now_str()
    else:
        db.session.add(ChapterVideo(chapter_no=chapter_no, chapter=chapter,
            student_class=sel_class, section=section,
            youtube_url=youtube_url, title=title, added_at=now_str()))
    cfg = ChapterConfig.query.filter_by(chapter_no=chapter_no, student_class=sel_class).first()
    if cfg: cfg.has_video = True
    db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/delete_video/<int:id>")
def delete_video(id):
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    v = db.session.get(ChapterVideo, id)
    sel_class = v.student_class if v else "6"
    if v: db.session.delete(v); db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/save_chapter_config", methods=["POST"])
def save_chapter_config():
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    chapter_no = int(request.form["chapter_no"])
    sel_class  = request.form.get("student_class","6")
    cfg = ChapterConfig.query.filter_by(chapter_no=chapter_no, student_class=sel_class).first()
    if not cfg:
        cfg = ChapterConfig(chapter_no=chapter_no, student_class=sel_class)
        db.session.add(cfg)
    cfg.is_active    = "is_active"    in request.form
    cfg.active_month = request.form.get("active_month","")
    cfg.has_video    = "has_video"    in request.form
    cfg.has_notes    = "has_notes"    in request.form
    cfg.has_book     = "has_book"     in request.form
    cfg.has_quiz     = "has_quiz"     in request.form
    cfg.has_game     = "has_game"     in request.form
    cfg.has_feedback = "has_feedback" in request.form
    db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/create_assignment", methods=["POST"])
def create_assignment():
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    filename  = ""
    sel_class = request.form.get("student_class","6")
    file = request.files.get("file")
    if file and file.filename:
        safe_name = safe_filename(file.filename)
        file_bytes = file.read()
        # Save to local disk (fast serving if still on disk)
        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
        with open(os.path.join(app.config["UPLOAD_FOLDER"], safe_name), "wb") as f:
            f.write(file_bytes)
        # Also save to Supabase so it survives server redeploys
        try:
            import mimetypes
            mime_type, _ = mimetypes.guess_type(safe_name)
            mime_type = mime_type or "application/octet-stream"
            supabase.storage.from_("lms-files").upload(
                safe_name,
                file_bytes,
                {"content-type": mime_type}
            )
        except Exception as e:
            print(f"Supabase backup upload warning: {e}")
        filename = safe_name
    db.session.add(Assignment(
        title=request.form["title"], chapter=request.form["chapter"],
        chapter_no=int(request.form.get("chapter_no",0)),
        student_class=sel_class, section=request.form["section"],
        description=request.form.get("description",""), due_date=request.form.get("due_date",""),
        filename=filename, created_at=now_str()))
    db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/delete_assignment/<int:id>")
def delete_assignment(id):

    if "teacher" not in session and "admin" not in session:
        return redirect("/teacher_login")

    assignment = Assignment.query.get_or_404(id)

    # delete related submissions first
    AssignmentSubmission.query.filter_by(
        assignment_id=id
    ).delete()

    db.session.delete(assignment)
    db.session.commit()

    return redirect("/teacher")

@app.route("/assignment_submissions/<int:id>")
def assignment_submissions(id):

    if "teacher" not in session and "admin" not in session:
        return redirect("/teacher_login")

    assignment = Assignment.query.get_or_404(id)

    # get all students of that class and section
    all_students = Student.query.filter_by(
        student_class=assignment.student_class,
        section=assignment.section
    ).all()


    # get submitted students
    submissions = AssignmentSubmission.query.filter_by(
        assignment_id=id
    ).all()


    sel_class = assignment.student_class


    return render_template(
        "assignment_submission.html",
        assignment=assignment,
        all_students=all_students,
        submissions=submissions,
        sel_class=sel_class
    )
def assignment_results(id):

    if "teacher" not in session:
        return redirect("/teacher_login")

    assignment = Assignment.query.get_or_404(id)

    submissions = AssignmentSubmission.query.filter_by(
        assignment_id=id
    ).all()

    return render_template(
        "assignment_results.html",
        assignment=assignment,
        submissions=submissions
    )
@app.route("/create_quiz", methods=["POST"])
def create_quiz():
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    chapter_no = int(request.form.get("chapter_no",0))
    sel_class  = request.form.get("student_class","6")
    quiz_type  = request.form.get("quiz_type","mcq")
    external_link = request.form.get("external_link","").strip()

    # FIX #4: validate external link quiz has a link before saving
    if quiz_type == "link" and not external_link:
        return redirect(f"/teacher?class={sel_class}&error=link_missing")

    quiz = Quiz(title=request.form["title"], chapter=request.form["chapter"],
        chapter_no=chapter_no, student_class=sel_class, section=request.form["section"],
        description=request.form.get("description",""), quiz_type=quiz_type,
        external_link=external_link, created_at=now_str())
    db.session.add(quiz); db.session.flush()
    if quiz_type == "mcq":
        questions = request.form.getlist("question_text")
        opts_a = request.form.getlist("option_a")
        opts_b = request.form.getlist("option_b")
        opts_c = request.form.getlist("option_c")
        opts_d = request.form.getlist("option_d")
        correct = request.form.getlist("correct_option")
        for i, qtext in enumerate(questions):
            if qtext.strip():
                db.session.add(QuizQuestion(
                    quiz_id=quiz.id, question_text=qtext.strip(),
                    option_a=opts_a[i] if i < len(opts_a) else "",
                    option_b=opts_b[i] if i < len(opts_b) else "",
                    option_c=opts_c[i] if i < len(opts_c) else "",
                    option_d=opts_d[i] if i < len(opts_d) else "",
                    correct_option=correct[i] if i < len(correct) else "A"))
    cfg = ChapterConfig.query.filter_by(chapter_no=chapter_no, student_class=sel_class).first()
    if cfg: cfg.has_quiz = True
    db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/delete_quiz/<int:id>")
def delete_quiz(id):
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    q = db.session.get(Quiz, id)
    sel_class = q.student_class if q else "6"
    if q: db.session.delete(q); db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/quiz_results/<int:quiz_id>")
def quiz_results(quiz_id):
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    quiz     = db.session.get(Quiz, quiz_id)
    attempts = QuizAttempt.query.filter_by(quiz_id=quiz_id).order_by(QuizAttempt.score.desc()).all()
    return render_template("quiz_results.html", quiz=quiz, attempts=attempts)

@app.route("/create_game", methods=["POST"])
def create_game():
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    chapter_no = int(request.form.get("chapter_no",0))
    sel_class  = request.form.get("student_class","6")
    sel_game_type = request.form.get("game_type", "matching")
    game = Game(title=request.form["title"], chapter=request.form["chapter"],
        chapter_no=chapter_no, student_class=sel_class,
        section=request.form["section"], created_at=now_str(),
        game_type=sel_game_type)
    db.session.add(game); db.session.flush()
    terms = request.form.getlist("term")
    definitions = request.form.getlist("definition")
    for t, d in zip(terms, definitions):
        t = t.strip(); d = d.strip()
        if sel_game_type == "word_cloud":
            # Sentence Typing game: only the sentence (term) is required
            if t:
                db.session.add(GamePair(game_id=game.id, term=t, definition=d))
        else:
            # Matching Pairs game: unchanged, both term and definition required
            if t and d:
                db.session.add(GamePair(game_id=game.id, term=t, definition=d))
    cfg = ChapterConfig.query.filter_by(chapter_no=chapter_no, student_class=sel_class).first()
    if cfg: cfg.has_game = True
    db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/delete_game/<int:id>")
def delete_game(id):
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    g = db.session.get(Game, id)
    sel_class = g.student_class if g else "6"
    if g: db.session.delete(g); db.session.commit()
    return redirect(f"/teacher?class={sel_class}")

@app.route("/game_leaderboard/<int:game_id>")
def game_leaderboard(game_id):
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    game   = db.session.get(Game, game_id)
    scores = GameScore.query.filter_by(game_id=game_id).order_by(GameScore.score.desc(), GameScore.time_seconds).all()
    return render_template("game_leaderboard.html", game=game, scores=scores)

# FIX #5/#9: view_feedbacks now shows queries with chapter filter and reply button
@app.route("/view_feedbacks")
def view_feedbacks():
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    sel_class  = request.args.get("class","6")
    chapter_no = request.args.get("chapter_no", 0, type=int)
    chapters   = ALL_CHAPTERS.get(sel_class, ALL_CHAPTERS["6"])
    query = StudentQuery.query.filter_by(student_class=sel_class)
    if chapter_no: query = query.filter_by(chapter_no=chapter_no)
    queries = query.order_by(StudentQuery.id.desc()).all()
    return render_template("view_feedbacks.html", queries=queries, chapters=chapters,
        selected_chapter=chapter_no, sel_class=sel_class, all_classes=["6","7","8","9"])

# FIX #9: teacher reply to query
@app.route("/reply_query/<int:query_id>", methods=["POST"])
def reply_query(query_id):
    if "teacher" not in session and "admin" not in session: return redirect("/teacher_login")
    q = db.session.get(StudentQuery, query_id)
    if q:
        q.teacher_reply = request.form.get("reply","").strip()
        q.replied_at = now_str()
        db.session.commit()
    sel_class = q.student_class if q else "6"
    return redirect(f"/view_feedbacks?class={sel_class}")

# ─────────────────────────────────────────────
# STUDENT PORTAL
# ─────────────────────────────────────────────

@app.route("/student")
def student_portal():
    if "student_id" not in session:
        return render_template("student.html", student=None, chapters=[],
            notes=[], books=[], quizzes=[], assignments=[], games=[], videos=[],
            progress_map={}, submitted_ids={}, attempted_quiz_ids={},
            game_scores={}, queries_done=set(), chapter_configs={},
            chapter_pcts={}, completed=0, in_progress=0, total_chapters=0,
            student_class="")

    student = db.session.get(Student, session["student_id"])
    if not student:
        session.pop("student_id", None)
        return redirect("/student")

    # ── Serve cached dashboard data if fresh (avoids 8 DB queries per load) ──
    cache_key = _student_cache_key(student.id)
    cached = cache.get(cache_key)
    if cached:
        return render_template("student.html", **cached)

    cls  = student.student_class
    sec  = student.section
    CHAPTERS = ALL_CHAPTERS.get(cls, ALL_CHAPTERS["6"])
    configs  = {c.chapter_no: c for c in ChapterConfig.query.filter_by(student_class=cls).all()}

    def mat(ftype):
        rows = Note.query.filter(Note.file_type==ftype, Note.student_class==cls) \
            .order_by(Note.chapter_no, Note.id).all()
        return [n for n in rows if note_visible_to_section(n.section, sec)]

    notes       = mat("notes")
    books       = mat("book")
    quizzes     = Quiz.query.filter(Quiz.student_class==cls,
                    (Quiz.section==sec)|(Quiz.section=="All Sections")).order_by(Quiz.chapter_no).all()
    assignments = Assignment.query.filter(Assignment.student_class==cls,
                    (Assignment.section==sec)|(Assignment.section=="")|(Assignment.section=="All Sections")).order_by(Assignment.chapter_no).all()
    games       = Game.query.filter(Game.student_class==cls,
                    (Game.section==sec)|(Game.section=="All Sections")).order_by(Game.chapter_no).all()
    videos      = ChapterVideo.query.filter(ChapterVideo.student_class==cls,
                    (ChapterVideo.section==sec)|(ChapterVideo.section=="All Sections")).order_by(ChapterVideo.chapter_no).all()

    progs             = {p.chapter_no: p for p in ChapterProgress.query.filter_by(student_id=student.id, student_class=cls).all()}
    submitted_ids     = {s.assignment_id: s.status for s in AssignmentSubmission.query.filter_by(student_id=student.id).all()}
    attempted_quiz_ids= {a.quiz_id: a for a in QuizAttempt.query.filter_by(student_id=student.id).all()}
    game_scores       = {s.game_id: s for s in GameScore.query.filter_by(student_id=student.id).all()}
    queries_done      = {q.chapter_no for q in StudentQuery.query.filter_by(student_id=student.id, student_class=cls).all()}
    student_queries   = StudentQuery.query.filter_by(student_id=student.id, student_class=cls).order_by(StudentQuery.id.desc()).all()

    chapter_pcts = {}
    for ch in CHAPTERS:
        cno = ch["no"]
        cfg = configs.get(cno)
        prog= progs.get(cno)
        chapter_pcts[cno] = get_chapter_progress_pct(prog, cfg)

    completed   = sum(1 for p in chapter_pcts.values() if p == 100)
    in_progress = sum(1 for p in chapter_pcts.values() if 0 < p < 100)

    ctx = dict(
        student=student, chapters=CHAPTERS, notes=notes, books=books,
        quizzes=quizzes, assignments=assignments, games=games, videos=videos,
        progress_map=progs, submitted_ids=submitted_ids,
        attempted_quiz_ids=attempted_quiz_ids, game_scores=game_scores,
        queries_done=queries_done, student_queries=student_queries,
        chapter_configs=configs,
        chapter_pcts=chapter_pcts, completed=completed,
        in_progress=in_progress, total_chapters=len(CHAPTERS),
        student_class=cls)

    cache.set(cache_key, ctx, timeout=60)   # cache for 60 seconds
    return render_template("student.html", **ctx)

@app.route("/student_login", methods=["POST"])
def student_login():
    admission_no = request.form["admission_no"]
    dob = request.form.get("dob","").strip()
    student = Student.query.filter_by(admission_no=admission_no).first()
    if student and (not dob or student.dob == dob):
        session["student_id"] = student.id
        return redirect("/student")
    return render_template("student.html", student=None, error="Admission No not found.",
        chapters=[], notes=[], books=[], quizzes=[], assignments=[], games=[], videos=[],
        progress_map={}, submitted_ids={}, attempted_quiz_ids={}, game_scores={},
        queries_done=set(), student_queries=[], chapter_configs={}, chapter_pcts={},
        completed=0, in_progress=0, total_chapters=0, student_class="")

@app.route("/student_logout")
def student_logout():
    session.pop("student_id", None)
    return redirect("/student")

# FIX #2: mark_video now uses POST to avoid "method not allowed" and correctly marks progress
@app.route("/mark_video/<int:chapter_no>", methods=["GET","POST"])
def mark_video(chapter_no):
    if "student_id" not in session: return redirect("/student")
    student = db.session.get(Student, session["student_id"])
    cls = student.student_class if student else "6"
    prog = ChapterProgress.query.filter_by(student_id=session["student_id"], chapter_no=chapter_no, student_class=cls).first()
    if not prog:
        prog = ChapterProgress(student_id=session["student_id"], chapter_no=chapter_no, student_class=cls)
        db.session.add(prog)
    prog.video_done = True; prog.updated_at = now_str()
    db.session.commit()
    invalidate_student_cache(session["student_id"])
    return redirect("/student")

# FIX #7: mark_notes accepts both GET and POST (was returning 405)
@app.route("/mark_notes/<int:chapter_no>", methods=["GET","POST"])
def mark_notes(chapter_no):
    if "student_id" not in session: return redirect("/student")
    student = db.session.get(Student, session["student_id"])
    cls = student.student_class if student else "6"
    prog = ChapterProgress.query.filter_by(student_id=session["student_id"], chapter_no=chapter_no, student_class=cls).first()
    if not prog:
        prog = ChapterProgress(student_id=session["student_id"], chapter_no=chapter_no, student_class=cls)
        db.session.add(prog)
    prog.notes_done = True; prog.updated_at = now_str()
    db.session.commit()
    invalidate_student_cache(session["student_id"])
    return redirect("/student")

@app.route("/mark_book/<int:chapter_no>", methods=["GET","POST"])
def mark_book(chapter_no):
    if "student_id" not in session: return redirect("/student")
    student = db.session.get(Student, session["student_id"])
    cls = student.student_class if student else "6"
    prog = ChapterProgress.query.filter_by(student_id=session["student_id"], chapter_no=chapter_no, student_class=cls).first()
    if not prog:
        prog = ChapterProgress(student_id=session["student_id"], chapter_no=chapter_no, student_class=cls)
        db.session.add(prog)
    prog.book_done = True; prog.updated_at = now_str()
    db.session.commit()
    invalidate_student_cache(session["student_id"])
    return redirect("/student")

@app.route("/submit_assignment/<int:assignment_id>")
def submit_assignment(assignment_id):
    if "student_id" not in session: return redirect("/student")
    if not AssignmentSubmission.query.filter_by(student_id=session["student_id"], assignment_id=assignment_id).first():
        db.session.add(AssignmentSubmission(student_id=session["student_id"],
            assignment_id=assignment_id, status="submitted", submitted_at=now_str()))
        db.session.commit()
        invalidate_student_cache(session["student_id"])
    return redirect("/student")

# FIX #8: attempt_quiz uses proper standalone template (no base.html dependency)
@app.route("/attempt_quiz/<int:quiz_id>", methods=["GET","POST"])
def attempt_quiz(quiz_id):
    if "student_id" not in session: return redirect("/student")
    quiz = db.session.get(Quiz, quiz_id)
    if not quiz: return redirect("/student")
    if quiz.quiz_type == "link":
        # Mark quiz done for external links
        student = db.session.get(Student, session["student_id"])
        cls = student.student_class if student else "6"
        prog = ChapterProgress.query.filter_by(student_id=session["student_id"], chapter_no=quiz.chapter_no, student_class=cls).first()
        if not prog:
            prog = ChapterProgress(student_id=session["student_id"], chapter_no=quiz.chapter_no, student_class=cls)
            db.session.add(prog)
        prog.quiz_done = True; prog.updated_at = now_str()
        db.session.commit()
        return redirect(quiz.external_link)
    existing = QuizAttempt.query.filter_by(student_id=session["student_id"], quiz_id=quiz_id).first()
    if request.method == "POST":
        score = sum(1 for q in quiz.questions if request.form.get(f"q_{q.id}","").upper() == q.correct_option.upper())
        total = len(quiz.questions)
        db.session.add(QuizAttempt(student_id=session["student_id"], quiz_id=quiz_id,
            score=score, total=total, attempted_at=now_str()))
        student = db.session.get(Student, session["student_id"])
        cls = student.student_class if student else "6"
        prog = ChapterProgress.query.filter_by(student_id=session["student_id"], chapter_no=quiz.chapter_no, student_class=cls).first()
        if not prog:
            prog = ChapterProgress(student_id=session["student_id"], chapter_no=quiz.chapter_no, student_class=cls)
            db.session.add(prog)
        prog.quiz_done = True; prog.updated_at = now_str()
        db.session.commit()
        invalidate_student_cache(session["student_id"])
        return render_template("quiz_score.html", score=score, total=total, quiz=quiz)
    return render_template("attempt_quiz.html", quiz=quiz)

# FIX #8: play_game — game template is standalone, no base.html
@app.route("/play_game/<int:game_id>")
def play_game(game_id):
    if "student_id" not in session: return redirect("/student")
    game = db.session.get(Game, game_id)
    if not game: return redirect("/student")
    existing    = GameScore.query.filter_by(student_id=session["student_id"], game_id=game_id).first()
    leaderboard = GameScore.query.filter_by(game_id=game_id).order_by(GameScore.score.desc(), GameScore.time_seconds).limit(10).all()
    pairs_data  = [p.to_dict() for p in game.pairs]
    template = "play_word_cloud.html" if getattr(game, "game_type", "matching") == "word_cloud" else "play_game.html"
    return render_template(template, game=game, existing=existing, leaderboard=leaderboard, pairs_data=pairs_data)

@app.route("/save_game_score", methods=["POST"])
def save_game_score():
    if "student_id" not in session: return jsonify({"ok": False})
    data         = request.get_json()
    game_id      = data.get("game_id")
    score        = data.get("score", 0)
    time_seconds = data.get("time_seconds", 999)
    game = db.session.get(Game, game_id)
    if not game: return jsonify({"ok": False})
    existing = GameScore.query.filter_by(student_id=session["student_id"], game_id=game_id).first()
    if existing:
        if score > existing.score or (score == existing.score and time_seconds < existing.time_seconds):
            existing.score = score; existing.time_seconds = time_seconds; existing.played_at = now_str()
    else:
        db.session.add(GameScore(student_id=session["student_id"], game_id=game_id,
            score=score, time_seconds=time_seconds, played_at=now_str()))
    student = db.session.get(Student, session["student_id"])
    cls = student.student_class if student else "6"
    prog = ChapterProgress.query.filter_by(student_id=session["student_id"], chapter_no=game.chapter_no, student_class=cls).first()
    if not prog:
        prog = ChapterProgress(student_id=session["student_id"], chapter_no=game.chapter_no, student_class=cls)
        db.session.add(prog)
    prog.game_done = True; prog.updated_at = now_str()
    db.session.commit()
    invalidate_student_cache(session["student_id"])
    leaderboard = []
    for s in GameScore.query.filter_by(game_id=game_id).order_by(GameScore.score.desc(), GameScore.time_seconds).limit(10).all():
        leaderboard.append({"name": s.student.student_name, "score": s.score, "time": s.time_seconds})
    return jsonify({"ok": True, "leaderboard": leaderboard})

# FIX #9: submit_query (replaces submit_feedback — not a progress item)
@app.route("/submit_query", methods=["POST"])
def submit_query():
    if "student_id" not in session: return redirect("/student")
    chapter_no = int(request.form["chapter_no"])
    student    = db.session.get(Student, session["student_id"])
    cls        = student.student_class if student else "6"
    db.session.add(StudentQuery(
        student_id=session["student_id"], chapter_no=chapter_no,
        student_class=cls,
        comment=request.form.get("comment",""),
        submitted_at=now_str()))
    db.session.commit()
    invalidate_student_cache(session["student_id"])
    return redirect("/student")

# Keep old feedback endpoint for backward compatibility
@app.route("/submit_feedback", methods=["POST"])
def submit_feedback():
    return submit_query()

# ── Read-only viewer ────────────────────────────────────────────────────────
# Wraps a file in an embedded viewer (PDF toolbar hidden, right-click/save
# shortcuts blocked) instead of letting students hit the raw file directly.
# Inlined as a string (rather than templates/view_file.html) so this whole
# feature lives in app.py alone — nothing extra to forget when deploying.
OFFICE_EXTS = {"doc", "docx", "ppt", "pptx", "xls", "xlsx"}

VIEW_FILE_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{{ title }}</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
  * { box-sizing: border-box; }
  html, body { margin:0; padding:0; height:100%; background:#2b2b2b; font-family:'Segoe UI',Arial,sans-serif; }
  .topbar {
    height:48px; background:#1f2937; color:#fff; display:flex; align-items:center;
    padding:0 16px; font-size:14px; font-weight:600; letter-spacing:.3px;
    justify-content:space-between; user-select:none;
  }
  .topbar .lbl { display:flex; align-items:center; gap:8px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
  .topbar .tag { background:#e74c3c; font-size:11px; padding:3px 8px; border-radius:10px; font-weight:700; }
  .frame-wrap { position:relative; width:100%; height:calc(100% - 48px); }
  iframe {
    width:100%; height:100%; border:none; background:#525659;
  }
  /* transparent shield along the very top of the iframe to block the
     native PDF viewer's own download/print icons in most browsers */
  .shield-top {
    position:absolute; top:0; left:0; right:0; height:40px; z-index:5;
    background:transparent;
  }
</style>
</head>
<body oncontextmenu="return false;">
  <div class="topbar">
    <div class="lbl">📖 {{ title }}</div>
    <div class="tag">READ ONLY</div>
  </div>
  <div class="frame-wrap">
    <div class="shield-top" title="Viewing only"></div>
    <iframe src="{{ file_url }}{% if viewer_mode == 'pdf' %}#toolbar=0&navpanes=0&scrollbar=1{% endif %}" title="{{ title }}"></iframe>
  </div>

  <script>
    // Block common save/print/devtools shortcuts. This is a deterrent for
    // casual users, not a hard security guarantee — a determined user can
    // always find a way to copy content they can see on screen.
    document.addEventListener('keydown', function (e) {
      const k = e.key ? e.key.toLowerCase() : '';
      const blockCombo =
        ((e.ctrlKey || e.metaKey) && (k === 's' || k === 'p' || k === 'u')) ||
        (e.ctrlKey && e.shiftKey && k === 'i') ||
        k === 'f12';
      if (blockCombo) { e.preventDefault(); e.stopPropagation(); return false; }
    });
    document.addEventListener('contextmenu', function (e) { e.preventDefault(); });
  </script>
</body>
</html>
"""

@app.route("/view/<path:filename>")
def view_file(filename):
    if "student_id" not in session and "teacher" not in session and "admin" not in session:
        return redirect("/student")
    title = request.args.get("title", filename)

    if filename.startswith("http://") or filename.startswith("https://"):
        file_url = to_embeddable_url(filename)
        viewer_mode = "direct"
    else:
        # Absolute URL — Office Online / Google Docs viewers must be able
        # to fetch the file themselves, a relative /uploads/... path won't work.
        file_url = request.host_url.rstrip("/") + f"/uploads/{filename}"
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext in OFFICE_EXTS:
            # Browsers have no built-in renderer for Word/Excel/PowerPoint —
            # loading them straight in an iframe just triggers a download.
            # Route through Microsoft's Office Online viewer instead, which
            # renders the document as a read-only preview.
            from urllib.parse import quote
            file_url = "https://view.officeapps.live.com/op/embed.aspx?src=" + quote(file_url, safe="")
            viewer_mode = "direct"
        else:
            viewer_mode = "pdf"  # pdf, images, text — native iframe rendering works

    return render_template_string(VIEW_FILE_HTML, file_url=file_url, title=title, viewer_mode=viewer_mode)

# ── File serving ──────────────────────────────────────────────────────────────
# All files (notes, books, assignments) are stored in Supabase storage.
# We redirect directly to the public CDN URL — no proxying through Flask.
# Make sure lms-files bucket is set to PUBLIC in Supabase Storage settings.
@app.route("/download/<path:filename>")
def download_file(filename):
    # Used only by the assignments section so students can save/print a copy.
    # Notes and books intentionally do NOT use this route - they stay
    # read-only via /view/<filename>.
    if "student_id" not in session and "teacher" not in session and "admin" not in session:
        return redirect("/student")

    import re
    # Assignment filenames are stored as "YYYYMMDD_HHMMSS_originalname.ext" -
    # strip that prefix so the downloaded file has a clean, readable name.
    display_name = re.sub(r'^\d{8}_\d{6}_', '', filename)

    local_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if os.path.exists(local_path):
        return send_from_directory(
            app.config["UPLOAD_FOLDER"], filename,
            as_attachment=True, download_name=display_name
        )

    # Not on local disk -> fetch from Supabase storage and stream back
    # with attachment headers (a plain redirect can't force a download).
    import urllib.request
    public_url = f"{SUPABASE_URL}/storage/v1/object/public/lms-files/{filename}"
    try:
        with urllib.request.urlopen(public_url) as resp:
            data = resp.read()
            mime = resp.headers.get_content_type()
    except Exception:
        return redirect("/student")

    from flask import Response
    response = Response(data, mimetype=mime or "application/octet-stream")
    response.headers["Content-Disposition"] = f'attachment; filename="{display_name}"'
    return response

@app.route("/uploads/<filename>")
def serve_upload(filename):
    local_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)

    # Try local disk first (assignments saved locally on this server)
    if os.path.exists(local_path):
        import mimetypes
        mime, _ = mimetypes.guess_type(filename)
        from flask import make_response
        response = make_response(send_from_directory(app.config["UPLOAD_FOLDER"], filename))
        response.headers["Content-Disposition"] = f"inline; filename={filename}"
        response.headers["Content-Type"] = mime or "application/octet-stream"
        return response

    # Not on local disk → redirect to Supabase public CDN
    # Works for notes, books, AND assignments uploaded to Supabase
    public_url = f"{SUPABASE_URL}/storage/v1/object/public/lms-files/{filename}"
    return redirect(public_url)

# ─────────────────────────────────────────────
# DB INIT
# ─────────────────────────────────────────────
with app.app_context():
    os.makedirs("uploads", exist_ok=True)
    db.create_all()
    # Migration: add game_type column if it does not exist yet (PostgreSQL safe)
    try:
        db.session.execute(db.text("ALTER TABLE game ADD COLUMN game_type VARCHAR(30) DEFAULT 'matching'"))
        db.session.commit()
    except Exception:
        db.session.rollback()  # column already exists — safe to ignore
    if Admin.query.count() == 0:
        db.session.add(Admin(username="admin", password="1234")); db.session.commit()
    if Teacher.query.count() == 0:
        db.session.add(Teacher(username="teacher", password="1234", name="Computer Teacher")); db.session.commit()
    for cls, chapters in ALL_CHAPTERS.items():
        for ch in chapters:
            if not ChapterConfig.query.filter_by(chapter_no=ch["no"], student_class=cls).first():
                db.session.add(ChapterConfig(chapter_no=ch["no"], student_class=cls))
    db.session.commit()

if __name__ == "__main__":
    # ⚠️  Development only. In production, run with gunicorn:
    # gunicorn --workers 4 --threads 4 --timeout 60 -b 0.0.0.0:5000 app:app
    app.run(debug=False)
